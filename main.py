import time
from generator import LCG
from shop import Shop
from product import Product
#from client import Client
from openpyxl import load_workbook

def printProducts(products):
    # Petla po liscie products do wyswietlania
    for product in products:
        # print(product.value1, product.value2, product.value3) - gdyby nie bylo funkcji __str__ w klasie
        print(product)
        
def saveProducts(products, column, month):
    # zapis zebranych danych z pojedynczego dnia do pliku
    wb = load_workbook(r"C:\Users\Kinga\Desktop\niduc\__pycache__\niduc_projekt.xlsx")
    sheet = wb[month]
    row = 7

    for product in products:
        sheet.cell(row, column).value = product.value1
        column = column+1
        sheet.cell(row, column).value = product.value2
        column = column+1
        sheet.cell(row, column).value = product.value3
        column = column-2
        row = row+1
            
    wb.save("niduc_projekt.xlsx")
    
def oneDayData(products):
    # Kod generuje symulowane dane dla trzech zdarzeń: wyswietlania, dodania do koszyka i zakupu
    # Zdarzenia, ktore moga wystapic:

    # 0 - Nie wybrano żadnej opcji
    # 1 - Wyświetlenie
    # 2 - Wyświetlenie i dodanie do koszyka
    # 3 - Wyświetlenie, dodanie do koszyka i zakup

    generator = LCG()
    shop = Shop()
    start_time = time.time()
    duration = 1  # czas trwania petli w sekundach
    errors = 0
    # Petla powtarzajaca sie tyle czasu ile zadano, w tym przypadku 3s:
    while (time.time() - start_time) < duration:

        # Losowanie ilosci klientow w danej chwili na stronie
        # nnumClients = generator.rand(0, 500)

        # Wyswietlenie produktu
        # Losowanie indeksu w tablicy produktow - zakres wiekszy  o 100 niz ilosc elementow w tablicy po to aby nie zawsze udalo sie wylosowac produkt
        los = generator.rand(0, len(products) + 100000)
        if (los < len(products)):
            products[los].value1 = products[los].value1 + 1

        # Wyswietlenie i dodanie produktu do koszyka
        los = generator.rand(0, len(products) + 1000000)
        if(los < len(products) and products[los].value2<products[los].value1):
            products[los].value2 = products[los].value2 + 1

        # Wyswietlenie i zakup
        los = generator.rand(0, len(products) + 1000000)
        if (los < len(products) and products[los].value3 < products[los].value2):
            products[los].value3 = products[los].value3 + 1

        # Liczba klientow w danej chwili = liczba wyswietlen produktow w danej chwili
        #for product in products:
        #    shop.clients = shop.clients + product.value1

        # Wylosowanie błędu serwera
        # Zakładamy, że mozliwych błędów będzie 10 - określają je liczby od 50 do 55 po kolei
        shop.error = generator.rand(0,10000000)           #Zakres 10^9 dac

        if(shop.error >=50 and shop.error <=55):
            errors=errors+1

        # Sprawdzanie warunkow przez caly czas dzialania sklepu
        checkConditions(products, shop)
        
def createData(products):
    for i in range (6,34,3):
        oneDayData(products)
        saveProducts(products, i, "listopad")
'''
    for j in range (6,97,3):
        oneDayData(products)
        saveProducts(products, j, "grudzien")
    
    for k in range (6,97,3):
        oneDayData(products)
        saveProducts(products, k, "styczen")

    for l in range (6,67,3):
        oneDayData(products)
        saveProducts(products, l, "luty")'''

# WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI
# WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI WARUNKI
def condition1(product):
    if (product.value3 >= 990):
        #print("Produkt o indeksie " + str(product.index) + " zostal wyswietlony wiecej niz 900 razy")
        i=+5

def read_promo_file(filename=r'promo.txt'):
    # pobieranie danych promocyjnych z pliku txt
    data = []
    with open(filename, 'r') as f:
        for line in f:
            day, promo, chain = line.strip().split(';')
            promo = float(promo)
            chain = chain.strip()
            data.append((int(day), promo, chain))
    return data

def shopErrors(err, shop):

    # Blad 50 - spowalnia prace serwera o 5%
    if (shop.error == 50):
        shop.speed = shop.speed - 5

    if (shop.error == 51):
        print("Wystapil blad 51")

    if (shop.error == 52):
        print("Wystapil blad 52")

    if (shop.error == 53):
        print("Wystapil blad 53")

    if (shop.error == 54):
        print("Wystapil blad 54")

    # Awaria jednego z serwerow!!!
    if (shop.error == 55):
        shop.runServers = shop.runServers - 1

def checkConditions(products, shop):
    # Sprawdzanie po kolei kazdego warunku

    # Sprawdzenie warunkow dla klasy Product
    # Funkcja enumerate() tworzy iterator z pary (index, wartosc) - uzyta do uzyskania indeksu elementu w tablicy
    for i, prod in enumerate(products):
        condition1(prod)

    # Sprawdzenie wystapionych bledow:
    shopErrors(shop.error, shop)

def main():
    num_samples = 2  # liczba próbek, ile chcemy wygenerowac
    # Lista produktow
    products = []
    i=0
    errors = 0
    row = 7
    
    # Przygotowanie arkusza danych do odczytu
    wb = load_workbook(r"C:\Users\Kinga\Desktop\niduc\__pycache__\niduc_projekt.xlsx")
    sheet = wb["listopad"]

    # Petla powtarzajaca sie tyle razy, ile probek chcemy wylosowac i uzupelniajaca wartosci
    for i in range(num_samples):
        # Pobranie indeksu produktu z bazy
        idx = sheet.cell(row, column=2).value
        price = sheet.cell(row, column = 4).value
        row = row +1
        
        # Stworzenie obiektu product
        product = Product(idx, price)
        products.append(product)  # Dodanie obiektu do tablicy produktów
        
    createData(products)
    print("Koniec petli czasowej")
    print("Bledy: " + str(errors))
    printProducts(products)

'''
    data =read_promo_file()
    print("==================")
    print(data)
'''

if __name__ == "__main__":
    main()